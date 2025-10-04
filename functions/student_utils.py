from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill
from functions.question_utils import QuestionUtils
from functions.page_utils import PageUtils
from functions.ocr_post_utils import OCRPostUtils
from functions.grade_utils import GradeUtils
from functions.excel_utils import ExcelUtils
from functions.chart_utils import ChartUtils
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
import re
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUTPUT_DIR = Path("processed_results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# --------------------------
# Extract questions & grades + page markers for a group of pages (one student)
# --------------------------
class StudentUtils:
    def extract_from_pages(pages: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, int], Dict[int, int]]:
        """
        Extract student_info, qmap, per_q_seen and page_markers map.
        page_markers: { printed_page_number: total_pages }
        NOTE: Now qmap entries include:
            - achieved_page (page where grade was seen)
            - q_crop: actual crop image file path for the question header
            - g_crop: actual crop image file path for the grade
        """
        student_info = {"Mat_num": None, "Seat_num": None}
        questions = []
        grades = []
        page_markers: Dict[int, int] = {}

        for page in pages:
            page_num = page["page"]
            for item in page.get("results", []):
                label = item.get("label")
                ocr_output = item.get("text")  # list of (text, conf)
                bbox = item.get("bbox", [0, 0, 0, 0])
                path = item.get("path")
                # print("path is ",path)
                # path = item.get("path") 
                # print(path) # path from YOLO detection

                if not ocr_output:
                    continue

                best_text, best_conf = max(ocr_output, key=lambda x: x[1])

                if label == "page_number":
                    pm = PageUtils.parse_page_marker(best_text)
                    if pm:
                        printed_cur, printed_total = pm
                        page_markers[printed_cur] = printed_total

                elif label == "Mat_num":
                    student_info["Mat_num"] = best_text.split(":")[-1].strip()

                elif label == "seat_num":
                    student_info["Seat_num"] = best_text.split(":")[-1].strip()

                elif label == "question_num":
                    qid, max_marks = QuestionUtils.parse_question_headline(best_text)
                    if qid:
                        questions.append({
                            "qnum": qid,
                            "max_marks": max_marks,
                            "page": page_num,
                            "y": bbox[1],
                            "q_crop": item.get("raw_path"),   # keep actual crop
                        })
                    

                elif label == "grades":
                    grades.append({
                        "text": best_text,
                        "conf": best_conf,
                        "page": page_num,
                        "y": (bbox[1] + bbox[3]) / 2,
                        "g_crop": item.get("raw_path"),   # keep actual crop
                    })
                

        # sort questions
        questions.sort(key=lambda q: (q["page"], q["y"]))

        qmap = {}
        per_q_seen = {}

        for i, q in enumerate(questions):
            qnum = q["qnum"]
            max_marks = q["max_marks"]
            page_start, y_start = q["page"], q["y"]

            if i + 1 < len(questions):
                page_end, y_end = questions[i + 1]["page"], questions[i + 1]["y"]
            else:
                page_end, y_end = float("inf"), float("inf")

            # candidate grades between this question and the next one
            candidate_grades = []
            for g in grades:
                if (g["page"] > page_start or (g["page"] == page_start and g["y"] >= y_start)) and \
                (g["page"] < page_end or (g["page"] == page_end and g["y"] < y_end)):
                    candidate_grades.append(g)

            per_q_seen[qnum] = len(candidate_grades)

            if candidate_grades:
                best_grade = max(candidate_grades, key=lambda g: g["conf"])
                qmap[qnum] = {
                    "raw": best_grade["text"],
                    "raw_conf": best_grade["conf"],
                    "max_marks": max_marks,
                    "page": page_start,
                    "achieved_page": best_grade["page"],
                    "q_crop": q.get("q_crop"),
                    "g_crop": best_grade.get("g_crop"),
                }
            else:
                qmap[qnum] = {
                    "raw": None,
                    "raw_conf": None,
                    "max_marks": max_marks,
                    "page": page_start,
                    "achieved_page": None,
                    "q_crop": q.get("q_crop"),
                    "g_crop": None,
                }
                    # Ensure missing question numbers are also included
        if qmap:
            max_q = max(int(k) for k in qmap.keys())
            for qid in range(1, max_q + 1):
                qid_str = str(qid)
                if qid_str not in qmap:
                    qmap[qid_str] = {
                        "raw": None,
                        "raw_conf": None,
                        "max_marks": None,
                        "page": None,
                        "achieved_page": None,
                        "q_crop": None,
                        "g_crop": None,
                        "error": "YOLO failed to detect"
                    }


        return student_info, qmap, per_q_seen, page_markers

    def build_student_row_and_flags(student_info: Dict[str, Any], qmap: Dict[str, Any], base_output_dir: Path,
                                    page_check_msg: str) -> Tuple[Dict[str, Any], Dict[str, bool], Dict[str, Optional[float]]]:
        """
        Returns (row, normalized_flags, numeric_achieved_per_q)
        Row includes Page_Check, Max_Total, Total_Achieved, Percent, Final_Mark
        Note: This function still produces a wide 'row' (legacy), but the combined excel creation now uses a pivoted layout.
        """
        def format_number_de(val: Optional[float], decimals: int = 2) -> str:
            """Format number with German decimal separator (comma)."""
            if val is None:
                return ""
            try:
                s = f"{val:.{decimals}f}"
                return s.replace(".", ",")
            except Exception:
                return str(val).replace(".", ",")

        # Use full names for clarity
        row = {
            "Matriculation Number": student_info.get("Mat_num"),
            "Seat Number": student_info.get("Seat_num"),
        }

        normalized_flags: Dict[str, bool] = {}
        numeric_achieved: Dict[str, Optional[float]] = {}
        total = 0.0

        for qnum in sorted(qmap.keys(), key=lambda x: int(x)):
            entry = qmap[qnum]
            raw = entry.get("raw")
            max_marks = entry.get("max_marks")
            page = entry.get("page")

            normalized, changed = OCRPostUtils.normalize_digits(raw) if raw is not None else (None, False)
            formatted_str, numeric_val = GradeUtils.format_grade_string_and_value(normalized, max_marks)

            # apply German decimal separator for grade strings
            if formatted_str:
                formatted_str = formatted_str.replace(".", ",")

            # create hyperlink to the page where the grade was found (if achieved_page present), but put in separate column in pivoted sheet
            detected_page_for_link = entry.get("achieved_page") or page
            detected_page_path = Path(base_output_dir) / f"image_{detected_page_for_link}" / "detected.jpg"
            link_target = str(detected_page_path) if detected_page_path.exists() else None
            # for legacy wide row (we'll keep the formatted string; hyperlinks moved to separate column in pivoted view)
            row[qnum] = formatted_str
            row[f"{qnum}_max"] = format_number_de(max_marks, 1) if max_marks is not None else ""

            if numeric_val is not None:
                total += numeric_val

            numeric_achieved[qnum] = numeric_val
            normalized_flags[qnum] = changed

        max_total = sum((entry.get("max_marks") or 0) for entry in qmap.values())

        row["Total_Achieved"] = format_number_de(total, 1)  # one digit after separator
        row["Max_Total"] = format_number_de(max_total, 0)
        row["Percent"] = format_number_de((total / max_total * 100) if max_total else 0.0, 1).replace(".", ",")
        row["Final_Mark"] = GradeUtils.map_percentage_to_mark((total / max_total * 100) if max_total else 0.0).replace(".", ",")
        row["Page_Check"] = page_check_msg

        return row, normalized_flags, numeric_achieved


    def save_students_excel_and_primus(
        students_rows: List[Dict[str, Any]],
        students_normalized_flags: List[Dict[str, bool]],
        students_numeric_per_q: List[Dict[str, Optional[float]]],
        students_qmap_list: List[Dict[str, Any]],
        students_per_q_seen: List[Dict[str, int]],
        output_path: Path,
        unique_id: str
    ):
        DEFAULT_GRADING_TABLE = [
            (0,  "5,0"),
            (50, "4,0"),
            (55, "3,7"),
            (60, "3,3"),
            (65, "3,0"),
            (70, "2,7"),
            (75, "2,3"),
            (80, "2,0"),
            (85, "1,7"),
            (90, "1,3"),
            (95, "1,0"),
        ]

        # Colors (hex without alpha for openpyxl)
        COLORS = {
            "full":        "C6EFCE",  # light green
            "zero":        "C00000",  # deep red (changed from FFC7CE)
            "no_match":    "EE82EE",  # violet
            "double_match":"800080",  # purple
            "invalid":     "FFA07A",  # orange-like for invalid/plausibility
            "normalized":  "FFA500",  # orange for normalized (distinct)
            "page_error":  "D3D3D3"   # light gray for page plausibility error
        }

        # Font color overrides for readability
        FONT_COLORS = {
            "double_match": "FFFFFF",  # white on purple
            "zero": "FFFFFF"  # white on deep red
        }
    
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(title="Question Overview")

        # Legends (unchanged)
        marks_legend = [
            ("full points", COLORS["full"]),
            ("zero", COLORS["zero"]),
            ("no match found", COLORS["no_match"]),
            ("double match found", COLORS["double_match"]),
            ("normalized (OCR correction)", COLORS["normalized"]),
        ]
        errors_legend = [
            ("page plausibility error", COLORS["page_error"]),
            ("invalid / plausibility", COLORS["invalid"]),
        ]

        start_row = 1
        ws.cell(row=start_row, column=1, value="Marks Legend").font = Font(bold=True)
        r = start_row + 1
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for label, color in marks_legend:
            ws.cell(row=r, column=1, value=label).border = border
            box = ws.cell(row=r, column=2, value="")
            box.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            box.border = border
            if color.upper() in [COLORS["double_match"].upper(), COLORS["zero"].upper()]:
                box.font = Font(color="FFFFFF")
            r += 1

        err_start_col = 4
        ws.cell(row=start_row, column=err_start_col, value="Errors Legend").font = Font(bold=True)
        r = start_row + 1
        for label, color in errors_legend:
            ws.cell(row=r, column=err_start_col, value=label).border = border
            box = ws.cell(row=r, column=err_start_col + 1, value="")
            box.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            box.border = border
            r += 1

        # ------------------------------
        # Table headers exactly as requested
        # ------------------------------
        table_start_row = start_row + max(len(marks_legend), len(errors_legend)) + 2
        headers = [
            "Matriculation Number",
            "Seat Number",
            "Question",
            "Q_Max",
            "Q_Page",
            "Question_Image",
            "Achieved",
            "Achieved_Image",
            "Achieved_Page",
            "Achieved_Relative (%)",
            "Error_Check",
            "Full_Page_Link",
        ]
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=table_start_row, column=ci, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # if h == "Error_Check":
            #     cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        images_to_embed: List[Tuple[str, str]] = []
        current_row = table_start_row + 1

        # Fill rows, one row per question instance (same as before but adapted to new columns)
        for row_dict, numeric_vals, qmap, per_q_seen, norm_flags in zip(
            students_rows, students_numeric_per_q, students_qmap_list, students_per_q_seen, students_normalized_flags
        ):
            # Step 0: get original values
            # Step 0: get original values
            mat = row_dict.get("Matriculation Number")
            seat = row_dict.get("Seat Number")

            # Step 1: normalize and track if changed
            mat_norm, mat_changed = OCRPostUtils.normalize_digits(mat) if mat else (mat, False)
            seat_norm, seat_changed = OCRPostUtils.normalize_digits(seat) if seat else (seat, False)

            # convert to strings and strip spaces
            mat_norm = str(mat_norm).strip() if mat_norm is not None else ""
            mat=mat_norm
            seat_norm = str(seat_norm).strip() if seat_norm is not None else ""
            seat=seat_norm

            # Step 2: write to Question Overview
            mat_cell = ws.cell(row=current_row, column=1, value=mat_norm)
            seat_cell = ws.cell(row=current_row, column=2, value=seat_norm)

    # Matriculation Number
            if mat_changed:
                mat_cell.fill = PatternFill(start_color=COLORS["normalized"], end_color=COLORS["normalized"], fill_type="solid")
            elif OCRPostUtils.classify_id_field(mat_norm) != "full":  # invalid
                mat_cell.fill = PatternFill(start_color=COLORS["invalid"], end_color=COLORS["invalid"], fill_type="solid")

            # Seat Number
            if seat_changed:
                seat_cell.fill = PatternFill(start_color=COLORS["normalized"], end_color=COLORS["normalized"], fill_type="solid")
            elif OCRPostUtils.classify_id_field(seat_norm) != "full":  # invalid
                seat_cell.fill = PatternFill(start_color=COLORS["invalid"], end_color=COLORS["invalid"], fill_type="solid")




            for qnum in sorted(qmap.keys(), key=lambda x: int(x)):
                entry = qmap[qnum]
                max_marks = entry.get("max_marks")
                q_page = entry.get("page")
                achieved_page = entry.get("achieved_page")
                q_crop = entry.get("q_crop")
                g_crop = entry.get("g_crop")

                achieved_num = numeric_vals.get(qnum)

                # Plausibility message
                status = QuestionUtils.classify_question_status(achieved_num, max_marks, per_q_seen.get(qnum, 0))
                if status == "no_match":
                    ec_msg = f"{qnum}: no grade detected"
                elif status == "double_match":
                    ec_msg = f"{qnum}: multiple grades detected ({per_q_seen[qnum]})"
                elif status == "invalid":
                    ec_msg = f"{qnum}: invalid achieved ({achieved_num}) vs max {max_marks}"
                else:
                    ec_msg = "OK"

                # Column map (A..L) => 1..12
                ws.cell(row=current_row, column=1, value=mat).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=current_row, column=2, value=seat).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=current_row, column=3, value=qnum).alignment = Alignment(horizontal="center", vertical="center") 

                # Q_Max (numeric if present)
                if max_marks is not None:
                    try:
                        max_val = float(max_marks)
                        c = ws.cell(row=current_row, column=4, value=max_val)
                        # Use one decimal place; Excel shows comma if opened in German locale.
                        c.number_format = "#,##0.0"
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        ws.cell(row=current_row, column=4, value=str(max_marks))
                else:
                    ws.cell(row=current_row, column=4, value="").alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=current_row, column=5, value=q_page).alignment = Alignment(horizontal="center", vertical="center")

                # Question_Image (F)
                if q_crop and Path(q_crop).exists():
                    images_to_embed.append((f"F{current_row}", q_crop))
                # place a text label for image presence (keeps cell non-empty; embedding done later)
                ws.cell(row=current_row, column=6, value="").alignment = Alignment(horizontal="center", vertical="center")

                # Achieved (numeric)
    # Achieved (numeric)
                achieved_cell = ws.cell(row=current_row, column=7, value=float(achieved_num) if achieved_num is not None else "")
                achieved_cell.alignment = Alignment(horizontal="center", vertical="center")
                if achieved_num is not None:
                    achieved_cell.number_format = "#,##0.0"
                else:
                    achieved_cell.value = ""

                # Apply fill color based on status
                status = QuestionUtils.classify_question_status(achieved_num, max_marks, per_q_seen.get(qnum, 0))
                if status == "full":
                    achieved_cell.fill = PatternFill(start_color=COLORS["full"], end_color=COLORS["full"], fill_type="solid")
                elif status == "zero":
                    achieved_cell.fill = PatternFill(start_color=COLORS["zero"], end_color=COLORS["zero"], fill_type="solid")
                elif status == "no_match":
                    achieved_cell.fill = PatternFill(start_color=COLORS["no_match"], end_color=COLORS["no_match"], fill_type="solid")
                elif status == "double_match":
                    achieved_cell.fill = PatternFill(start_color=COLORS["double_match"], end_color=COLORS["double_match"], fill_type="solid")
                elif norm_flags.get(qnum, False):  # normalized OCR correction
                    achieved_cell.fill = PatternFill(start_color=COLORS["normalized"], end_color=COLORS["normalized"], fill_type="solid")


                # Achieved_Page
                ws.cell(row=current_row, column=9, value=achieved_page or "").alignment = Alignment(horizontal="center", vertical="center")

                # Grade_Image (I)
                if g_crop and Path(g_crop).exists():
                    images_to_embed.append((f"H{current_row}", g_crop))
                ws.cell(row=current_row, column=8, value="").alignment = Alignment(horizontal="center", vertical="center")

                # Achieved_Relative (%) formula -> =IF(Drow=0,"",Grow/Drow)
                # Put as formula; format as percentage with one decimal place.
                rel_cell = ws.cell(row=current_row, column=10, value=f'=IF(D{current_row}=0,"",G{current_row}/D{current_row})')
                rel_cell.number_format = "0.0%"
                rel_cell.alignment = Alignment(horizontal="center", vertical="center")


                # Error_Check
                err_cell = ws.cell(row=current_row, column=11, value=ec_msg)
                err_cell.alignment = Alignment(horizontal="center", vertical="center")

                if ec_msg == "OK":
                    err_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
                elif "no grade detected" in ec_msg:
                    err_cell.fill = PatternFill(start_color=COLORS["no_match"], end_color=COLORS["no_match"], fill_type="solid")
                elif "multiple grades detected" in ec_msg:
                    err_cell.fill = PatternFill(start_color=COLORS["double_match"], end_color=COLORS["double_match"], fill_type="solid")
                elif "invalid achieved" in ec_msg:
                    err_cell.fill = PatternFill(start_color=COLORS["invalid"], end_color=COLORS["invalid"], fill_type="solid")
                elif "page" in ec_msg.lower():  # plausibility or page error
                    err_cell.fill = PatternFill(start_color=COLORS["page_error"], end_color=COLORS["page_error"], fill_type="solid")

                # Full_Page_Link (L)
                # Build a link to the detected page if available; otherwise link to page image name.
                page_file = OUTPUT_DIR / unique_id / f"image_{achieved_page or q_page}" / "page.jpg"
                link_formula = ExcelUtils.make_clickable_link(page_file, "Open Page")
                ws.cell(row=current_row, column=12, value=link_formula).alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1

        # Apply borders, autofilter, column widths
        table_end_row = current_row - 1
        if table_end_row >= table_start_row + 1:
            ws.auto_filter.ref = f"A{table_start_row}:L{table_end_row}"
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # iterate rows in range and apply border to header columns
            for row_cells in ws[table_start_row:table_end_row]:
                # row_cells gives tuple of cells starting at column A
                for idx, c in enumerate(row_cells[: len(headers)]):
                    c.border = border

        # Column widths (some reasonable defaults)
        try:
            ws.column_dimensions["A"].width = 20 # Matriculation Number
            ws.column_dimensions["B"].width = 15 # Seat Number
            ws.column_dimensions["C"].width = 12  # Question
            ws.column_dimensions["D"].width = 12   # Q_Max
            ws.column_dimensions["E"].width = 12   # Q_Page
            ws.column_dimensions["F"].width = 25  # Question_Image
            ws.column_dimensions["G"].width = 15  # Achieved
            ws.column_dimensions["H"].width = 25  # Grade_Image
            ws.column_dimensions["I"].width = 15  #Achieved_page
            ws.column_dimensions["J"].width = 18  # Achieved_Relative (%)
            ws.column_dimensions["K"].width = 28  # Error_Check
            ws.column_dimensions["L"].width = 22  # Full_Page_Link
        except Exception:
            pass

        # Embed images (question crops into F, grade crops into I)
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils import column_index_from_string

        for coord, path in images_to_embed:
            try:
                img = XLImage(path)
                max_width, max_height = 120, 90
                ratio = min(max_width / img.width, max_height / img.height)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)

                # --- calculate offsets for centering ---
                col_letters = re.sub(r"\d+", "", coord)
                row_num = int(re.sub(r"[A-Z]+", "", coord))
                col_num = column_index_from_string(col_letters)

                col_width = ws.column_dimensions[col_letters].width or 10
                # Excel column width to pixels (approx factor ~7)
                cell_w = int(col_width * 7)
                cell_h = int(ws.row_dimensions[row_num].height or 70)

                offset_x = max((cell_w - img.width) // 2, 0)
                offset_y = max((cell_h - img.height) // 2, 0)

                # --- build proper anchor with size ---
                marker = AnchorMarker(col=col_num - 1, colOff=offset_x * 9525,
                                    row=row_num - 1, rowOff=offset_y * 9525)
                ext = XDRPositiveSize2D(img.width * 9525, img.height * 9525)
                anchor = OneCellAnchor(_from=marker, ext=ext)

                img.anchor = anchor
                ws.add_image(img)

                # lock row height so images don’t overlap
                ws.row_dimensions[row_num].height = cell_h

            except Exception as e:
                print(f"Image embedding failed for {path}: {e}")


        # Save workbook to output_path (Question Overview done)
        wb.save(output_path)

    
    # --------------------------
    # Primus_Export sheet with German grading system
    # --------------------------
        wb2 = load_workbook(output_path)
        if "Primus_Export" in wb2.sheetnames:
            primus = wb2["Primus_Export"]
        else:
            primus = wb2.create_sheet(title="Primus_Export")

        # Clear existing content
        for row in primus["A:Z"]:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()
                cell.border = Border()




        legend_start_row = 1
        primus.cell(row=legend_start_row, column=1, value="Grading Table").font = Font(bold=True)

        r = legend_start_row + 1
        for bound, grade in DEFAULT_GRADING_TABLE:
            primus.cell(row=r, column=1, value=bound)
            primus.cell(row=r, column=2, value=grade)
            primus.cell(row=r, column=1).number_format = "0"
            r += 1

        # Apply borders to grading table
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row_cells in primus.iter_rows(min_row=legend_start_row, max_row=r-1, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add Pass/Fail legend
        primus.cell(row=legend_start_row, column=4, value="Legend").font = Font(bold=True)

        primus.cell(row=legend_start_row+1, column=4, value="Pass (<5,0)").border = border
        cell_pass = primus.cell(row=legend_start_row+1, column=5, value="")
        cell_pass.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell_pass.border = border

        primus.cell(row=legend_start_row+2, column=4, value="Fail (=5,0)").border = border
        cell_fail = primus.cell(row=legend_start_row+2, column=5, value="")
        cell_fail.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell_fail.border = border


        # ---------------------------
        # Data Table (headers + rows below legend)
        # ---------------------------
        primus_headers = [
            "Matriculation Number",
            "Seat Number",
            "Achieved Marks",
            "Maximum Marks",
            "Percentage",
            "Final Mark",
            "Error Checking"
        ]

        header_row = r + 2  # leave some space after grading table + legend
        for ci, h in enumerate(primus_headers, start=1):
            c = primus.cell(row=header_row, column=ci, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # if h == "Error Checking":
            #     c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        rownum = header_row + 1
        grading_table_start = legend_start_row + 1
        grading_table_end = r - 1
        grading_range = f"$A${grading_table_start}:$B${grading_table_end}"  # dynamic range

        for row_dict in students_rows:
            primus.cell(row=rownum, column=1, value=mat)
            primus.cell(row=rownum, column=2, value=seat)
            
            # Achieved Marks
            primus.cell(
                row=rownum,
                column=3,
                value=f"=SUMIFS('Question Overview'!G:G,'Question Overview'!A:A,A{rownum},'Question Overview'!B:B,B{rownum})",
            )
            # Maximum Marks
            primus.cell(
                row=rownum,
                column=4,
                value=f"=SUMIFS('Question Overview'!D:D,'Question Overview'!A:A,A{rownum},'Question Overview'!B:B,B{rownum})",
            )
            primus.cell(row=rownum, column=5, value=f"=IF(D{rownum}=0,0,C{rownum}/D{rownum})")
            # Final Mark using German grading system
            primus.cell(row=rownum, column=6, value=f"=VLOOKUP(E{rownum}*100,{grading_range},2,TRUE)")
            # Error Checking
            primus.cell(row=rownum, column=7, value=f'=IF(OR(C{rownum}<0,C{rownum}>D{rownum}),"Range Error","OK")')

            # Number formats
            try:
                primus.cell(row=rownum, column=3).number_format = "#,##0.0"  # Achieved Marks
                primus.cell(row=rownum, column=4).number_format = "#,##0.0"  # Maximum Marks
                primus.cell(row=rownum, column=5).number_format = "0.0%"     # Percentage
                primus.cell(row=rownum, column=6).number_format = "0.0"      # Final Mark
            except Exception:
                pass

            rownum += 1


        primus.auto_filter.ref = f"A{header_row}:G{rownum-1}"

        # ---------------------------
        # Conditional Formatting
        # ---------------------------
        # Error Checking column (H)
        ok_rule = FormulaRule(
            formula=['INDIRECT("G"&ROW())="OK"'],
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        )
        bad_rule = FormulaRule(
            formula=['INDIRECT("G"&ROW())<>"OK"'],
            fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        )
        primus.conditional_formatting.add(f"F{header_row+1}:G{rownum-1}", ok_rule)
        primus.conditional_formatting.add(f"F{header_row+1}:G{rownum-1}", bad_rule)

        # Final Mark column (G): Pass/Fail
        primus.conditional_formatting.add(
            f"F{header_row+1}:F{rownum-1}",
            CellIsRule(operator="equal", formula=['"5,0"'],
                    fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"))  # Fail
        )
        primus.conditional_formatting.add(
            f"F{header_row+1}:F{rownum-1}",
            CellIsRule(operator="lessThan", formula=['"5,0"'],
                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))  # Pass
        )

        # ---------------------------
        # Borders for data table
        # ---------------------------
        num_columns = len(primus_headers)
        for row in primus.iter_rows(min_row=header_row, max_row=rownum-1, min_col=1, max_col=num_columns):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        try:
            cols = ["A", "B", "C", "D", "E", "F", "G"]  # now 7 columns
            widths = [20, 15, 18, 18, 18, 14, 22]       # adjust widths for the 7 columns

            for col, w in zip(cols, widths):
                primus.column_dimensions[col].width = w
        except Exception:
            pass

        wb2.save(output_path)






