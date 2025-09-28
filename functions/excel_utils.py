import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Optional
from pathlib import Path

class ExcelUtils:
    def save_excel_with_highlighting(df: pd.DataFrame, all_flags, all_statuses, all_page_checks, path):
        COLORS = {
        "full":        "C6EFCE",
        "partial":     "FFEB9C",
        "zero":        "FFC7CE",
        "no_match":    "EE82EE",
        "double_match":"800080",
        "invalid":     "FFA07A",
        "normalized":  "FFA500",
        "page_error":  "D3D3D3"
        }
        df.to_excel(path, index=False)
        wb = load_workbook(path)
        ws = wb.active

        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            flags = all_flags[r_idx-2]
            statuses = all_statuses[r_idx-2]
            page_ok = all_page_checks[r_idx-2]

            for qnum, changed in flags.items():
                col_idx = df.columns.get_loc(qnum) + 1
                cell = ws.cell(row=r_idx, column=col_idx)
                status = statuses.get(qnum, "invalid")
                fill_color = COLORS.get(status)
                if not page_ok:
                    fill_color = COLORS["page_error"]
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if changed:
                    cell.fill = PatternFill(start_color=COLORS["normalized"], end_color=COLORS["normalized"], fill_type="solid")

        wb.save(path)

    def make_clickable_link(path: Optional[str], text: str) -> str:
        """
        Make Excel HYPERLINK formula pointing to absolute file:// URI for portability.
        """
        if not path:
            return text or ""
        p = Path(path).resolve()
        uri = p.as_uri()
        # return Excel formula
        return f'=HYPERLINK("{uri}", "{text}")'
    

